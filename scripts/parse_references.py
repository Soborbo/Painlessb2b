"""
Parse all reference XLSX and MD files from the recommercialdevelopmentplan.zip
and generate a single import_data.json for the B2B Prospect Tracker API.
"""

import openpyxl
import json
import re
import os
import sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.join(os.environ.get('TEMP', '/tmp'), 'b2b-ref')
OUTPUT = os.path.join(os.path.dirname(__file__), 'import_data.json')

prospects = []
seen_names = set()


def normalize_name(name):
    """Normalize company name for dedup."""
    if not name:
        return ''
    return re.sub(r'\s+', ' ', str(name).strip().lower())


def clean(val):
    """Clean cell value."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('—', '–', '-', '', 'None', 'Via website', 'Via agents', 'TBC'):
        return None
    return s


def map_priority(val):
    """Map priority from various formats."""
    if not val:
        return 'medium'
    s = str(val).upper().strip()
    if 'HIGH' in s or '★★★' in s:
        return 'high'
    if 'LOW' in s or '★' == s:
        return 'low'
    return 'medium'


def map_status(val):
    """Map pipeline stage to app status."""
    if not val:
        return 'new'
    s = str(val).lower().strip()
    if 'not contacted' in s or 'not started' in s:
        return 'new'
    if 'contacted' in s or 'initial contact' in s:
        return 'contacted'
    if 'meeting' in s or 'in conversation' in s or 'negotiat' in s:
        return 'in_conversation'
    if 'follow' in s:
        return 'follow_up'
    if 'partner' in s or 'won' in s or 'active' in s:
        return 'partner'
    if 'rejected' in s or 'lost' in s:
        return 'rejected'
    if 'not interested' in s:
        return 'not_interested'
    return 'new'


def extract_email(val):
    """Extract email from a cell that may contain LinkedIn or other text."""
    if not val:
        return None
    s = str(val)
    match = re.search(r'[\w.+-]+@[\w.-]+\.\w+', s)
    return match.group(0) if match else None


def extract_website(val):
    """Clean website URL."""
    if not val:
        return None
    s = str(val).strip()
    if s.startswith('http'):
        return s
    if '.' in s and ' ' not in s:
        return f'https://{s}'
    return None


def add_prospect(name, category, address=None, postcode=None, phone=None,
                 website=None, generic_email=None, contact_name=None,
                 contact_email=None, status='new', priority='medium',
                 source='deep_research', notes=None):
    """Add prospect if not duplicate."""
    if not name:
        return
    norm = normalize_name(name)
    if norm in seen_names:
        return
    seen_names.add(norm)

    prospect = {
        'name': str(name).strip(),
        'category': category,
        'status': status,
        'priority': priority,
        'source': source,
    }
    if address:
        prospect['address'] = address
    if postcode:
        prospect['postcode'] = postcode
    if phone:
        prospect['phone'] = phone
    if website:
        prospect['website'] = website
    if generic_email:
        prospect['generic_email'] = generic_email
    if contact_name:
        prospect['contact_name'] = contact_name
    if contact_email:
        prospect['contact_email'] = contact_email
    if notes:
        prospect['notes'] = notes
    prospects.append(prospect)


def parse_carehomes_xlsx():
    """Parse carehomes_painless_removals.xlsx — Master Pipeline sheet."""
    path = os.path.join(BASE, 'carehomes_painless_removals.xlsx')
    wb = openpyxl.load_workbook(path)

    # Master Pipeline sheet
    ws = wb['Master Pipeline']
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}

        name = clean(vals.get('Company/Group'))
        if not name:
            continue

        tier = clean(vals.get('Tier')) or ''
        comp_type = clean(vals.get('Type')) or ''
        sites = clean(vals.get('Sites in Area')) or ''
        facilities = clean(vals.get('Key Facilities')) or ''
        locations = clean(vals.get('Location(s)')) or ''
        website = extract_website(vals.get('Website'))
        contact = clean(vals.get('Decision Maker'))
        title = clean(vals.get('Title')) or ''
        phone = clean(vals.get('Phone'))
        email_raw = clean(vals.get('Email/LinkedIn'))
        stage = clean(vals.get('Pipeline Stage'))
        priority = clean(vals.get('Priority'))
        raw_notes = clean(vals.get('Notes/Intelligence')) or ''

        # Determine category based on type
        category = 'Care Home'
        if 'retirement' in comp_type.lower() or 'village' in comp_type.lower():
            category = 'Retirement Village'
        elif 'charity' in comp_type.lower() and 'ha' in comp_type.lower():
            category = 'Housing Association'

        contact_email = extract_email(email_raw)
        generic_email = None
        # Only assign email if it looks like a real address (not a pattern)
        if contact_email and ('firstname' in contact_email.lower() or 'lastname' in contact_email.lower()):
            # This is an email pattern, not real - store in notes instead
            raw_notes = f'Email pattern: {contact_email}. {raw_notes}' if raw_notes else f'Email pattern: {contact_email}'
            contact_email = None
        elif contact_email and not contact:
            generic_email = contact_email
            contact_email = None

        notes_parts = []
        if tier:
            notes_parts.append(f'Tier: {tier}')
        if comp_type:
            notes_parts.append(f'Type: {comp_type}')
        if sites:
            notes_parts.append(f'Sites in area: {sites}')
        if facilities:
            notes_parts.append(f'Key facilities: {facilities}')
        if title:
            notes_parts.append(f'Contact title: {title}')
        if raw_notes:
            notes_parts.append(raw_notes)

        add_prospect(
            name=name,
            category=category,
            address=locations,
            phone=phone,
            website=website,
            generic_email=generic_email,
            contact_name=contact,
            contact_email=contact_email,
            status=map_status(stage),
            priority=map_priority(priority),
            notes='. '.join(notes_parts) if notes_parts else None,
        )

    # Also parse Outreach Tracker for additional contacts
    if 'Outreach Tracker' in wb.sheetnames:
        ws2 = wb['Outreach Tracker']
        headers2 = [str(c.value).strip() if c.value else '' for c in ws2[1]]
        for row in ws2.iter_rows(min_row=2, values_only=False):
            vals = {headers2[i]: row[i].value for i in range(len(headers2)) if i < len(row)}
            name = clean(vals.get('Company'))
            if not name:
                continue
            # If already added, skip (dedup handles it)
            norm = normalize_name(name)
            if norm in seen_names:
                continue
            contact = clean(vals.get('Contact Name'))
            priority = clean(vals.get('Priority'))
            raw_notes = clean(vals.get('Notes')) or ''
            add_prospect(
                name=name,
                category='Care Home',
                contact_name=contact,
                priority=map_priority(priority),
                notes=raw_notes if raw_notes else None,
            )


def parse_housing_associations_xlsx():
    """Parse Housing_associations.xlsx — Master Pipeline sheet."""
    path = os.path.join(BASE, 'Housing_associations.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Master Pipeline']
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}

        name = clean(vals.get('Organisation'))
        if not name:
            continue

        tier = clean(vals.get('Tier')) or ''
        org_type = clean(vals.get('Type')) or ''
        total_stock = clean(vals.get('Total Stock')) or ''
        older_stock = clean(vals.get('Older Persons Stock')) or ''
        schemes = clean(vals.get('Key Schemes in Area')) or ''
        coverage = clean(vals.get('Coverage')) or ''
        website = extract_website(vals.get('Website'))
        contact = clean(vals.get('Decision Maker'))
        title = clean(vals.get('Title')) or ''
        phone = clean(vals.get('Phone'))
        email_raw = clean(vals.get('Email'))
        procurement = clean(vals.get('Procurement Route')) or ''
        stage = clean(vals.get('Pipeline Stage'))
        priority = clean(vals.get('Priority'))
        raw_notes = clean(vals.get('Notes/Intelligence')) or ''

        contact_email = extract_email(email_raw)
        generic_email = None
        if email_raw and not contact_email:
            generic_email = email_raw
        elif contact_email and not contact:
            generic_email = contact_email
            contact_email = None

        notes_parts = []
        if tier:
            notes_parts.append(f'Tier: {tier}')
        if org_type:
            notes_parts.append(f'Type: {org_type}')
        if total_stock:
            notes_parts.append(f'Total stock: {total_stock}')
        if older_stock:
            notes_parts.append(f'Older persons stock: {older_stock}')
        if schemes:
            notes_parts.append(f'Key schemes: {schemes}')
        if title:
            notes_parts.append(f'Contact title: {title}')
        if procurement:
            notes_parts.append(f'Procurement: {procurement}')
        if raw_notes:
            notes_parts.append(raw_notes)

        add_prospect(
            name=name,
            category='Housing Association',
            address=coverage,
            phone=phone,
            website=website,
            generic_email=generic_email,
            contact_name=contact,
            contact_email=contact_email,
            status=map_status(stage),
            priority=map_priority(priority),
            notes='. '.join(notes_parts) if notes_parts else None,
        )


def parse_offices_xlsx():
    """Parse painless_removals_offices_pipeline.xlsx — Master Pipeline sheet."""
    path = os.path.join(BASE, 'painless_removals_offices_pipeline.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Master Pipeline']
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    # Category mapping by type
    type_to_category = {
        'property mgmt': 'Property Manager',
        'estate agent': 'Property Manager',
        'commercial landlord': 'Commercial Landlord',
        'developer': 'Commercial Landlord',
        'serviced office': 'Serviced Office',
        'coworking': 'Serviced Office',
        'creative coworking': 'Serviced Office',
        'innovation': 'Serviced Office',
        'business park': 'Business Park',
        'facilities management': 'FM Company',
        'fit-out': 'Fit-Out Contractor',
        'refurbishment': 'Fit-Out Contractor',
        'construction': 'Fit-Out Contractor',
    }

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}

        name = clean(vals.get('Organisation'))
        if not name:
            continue

        tier = clean(vals.get('Tier')) or ''
        org_type = clean(vals.get('Type')) or ''
        portfolio = clean(vals.get('Portfolio / Sites in Area')) or ''
        buildings = clean(vals.get('Key Buildings Managed/Owned')) or ''
        coverage = clean(vals.get('Coverage')) or ''
        website = extract_website(vals.get('Website'))
        contact = clean(vals.get('Decision Maker'))
        title = clean(vals.get('Title')) or ''
        phone = clean(vals.get('Phone'))
        email_raw = clean(vals.get('Email/Contact'))
        stage = clean(vals.get('Pipeline Stage'))
        priority = clean(vals.get('Priority'))
        raw_notes = clean(vals.get('Notes/Intelligence')) or ''

        # Determine category
        category = 'Other'
        type_lower = org_type.lower() if org_type else ''
        for key, cat in type_to_category.items():
            if key in type_lower:
                category = cat
                break

        contact_email = extract_email(email_raw)
        generic_email = None
        if contact_email and not contact:
            generic_email = contact_email
            contact_email = None

        notes_parts = []
        if tier:
            notes_parts.append(f'Tier: {tier}')
        if org_type:
            notes_parts.append(f'Type: {org_type}')
        if portfolio:
            notes_parts.append(f'Portfolio: {portfolio}')
        if buildings:
            notes_parts.append(f'Key buildings: {buildings}')
        if title:
            notes_parts.append(f'Contact title: {title}')
        if raw_notes:
            notes_parts.append(raw_notes)

        add_prospect(
            name=name,
            category=category,
            address=coverage,
            phone=phone,
            website=website,
            generic_email=generic_email,
            contact_name=contact,
            contact_email=contact_email,
            status=map_status(stage),
            priority=map_priority(priority),
            notes='. '.join(notes_parts) if notes_parts else None,
        )


def parse_estate_agents_xlsx():
    """Parse painless_removals_estate_agents_developers_pipeline.xlsx."""
    path = os.path.join(BASE, 'painless_removals_estate_agents_developers_pipeline.xlsx')
    wb = openpyxl.load_workbook(path)

    # Estate Agents sheet
    if 'Estate Agents' in wb.sheetnames:
        ws = wb['Estate Agents']
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=False):
            vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}

            name = clean(vals.get('Brand'))
            if not name:
                continue

            tier = clean(vals.get('Tier')) or ''
            parent = clean(vals.get('Parent Group')) or ''
            agent_type = clean(vals.get('Type')) or ''
            branches = clean(vals.get('Branches in Area')) or ''
            coverage = clean(vals.get('Coverage')) or ''
            phone = clean(vals.get('Phone'))
            website = extract_website(vals.get('Website'))
            contact = clean(vals.get('Contact / Decision Maker'))
            referral = clean(vals.get('Referral Model')) or ''
            priority = clean(vals.get('Priority'))
            raw_notes = clean(vals.get('Notes / Intelligence')) or ''

            notes_parts = []
            if tier:
                notes_parts.append(f'Tier: {tier}')
            if parent:
                notes_parts.append(f'Parent: {parent}')
            if agent_type:
                notes_parts.append(f'Type: {agent_type}')
            if branches:
                notes_parts.append(f'Branches in area: {branches}')
            if referral:
                notes_parts.append(f'Referral model: {referral}')
            if raw_notes:
                notes_parts.append(raw_notes)

            add_prospect(
                name=name,
                category='Estate Agent',
                address=coverage,
                phone=phone,
                website=website,
                contact_name=contact,
                priority=map_priority(priority),
                notes='. '.join(notes_parts) if notes_parts else None,
            )

    # Developers sheet
    if 'Developers & Developments' in wb.sheetnames:
        ws = wb['Developers & Developments']
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=False):
            vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}

            name = clean(vals.get('Developer'))
            if not name:
                continue

            parent = clean(vals.get('Parent')) or ''
            sites = clean(vals.get('Active Sites in Corridor')) or ''
            homes = clean(vals.get('Total Homes Planned')) or ''
            coverage = clean(vals.get('Coverage')) or ''
            phone = clean(vals.get('Sales Office Contact'))
            website = extract_website(vals.get('Website'))
            contact = clean(vals.get('New Homes Sales Team'))
            referral = clean(vals.get('Referral Model')) or ''
            priority = clean(vals.get('Priority'))
            raw_notes = clean(vals.get('Notes / Intelligence')) or ''

            notes_parts = []
            if parent:
                notes_parts.append(f'Parent: {parent}')
            if sites:
                notes_parts.append(f'Active sites: {sites}')
            if homes:
                notes_parts.append(f'Total homes planned: {homes}')
            if referral:
                notes_parts.append(f'Referral model: {referral}')
            if raw_notes:
                notes_parts.append(raw_notes)

            add_prospect(
                name=name,
                category='Developer',
                address=coverage,
                phone=phone,
                website=website,
                contact_name=contact,
                priority=map_priority(priority),
                notes='. '.join(notes_parts) if notes_parts else None,
            )


def parse_md_enrichment():
    """Parse MD files to enrich existing prospects with more detailed notes."""
    md_files = [
        'Later-life-prospect-pipeline.md',
        'housing-associations-pipeline.md',
        'commercial-office-prospects-pipeline.md',
    ]

    # Build lookup by normalized name
    prospect_lookup = {}
    for p in prospects:
        prospect_lookup[normalize_name(p['name'])] = p

    for fname in md_files:
        path = os.path.join(BASE, fname)
        if not os.path.exists(path):
            continue

        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Extract table rows — look for | # | or | N | patterns
        # Find companies mentioned in bold **Name** format in tables
        for match in re.finditer(r'\*\*([^*]+)\*\*', content):
            company = match.group(1).strip()
            norm = normalize_name(company)
            # Check if this company is already in our prospects
            if norm in prospect_lookup:
                # Get surrounding context (200 chars around the match)
                start = max(0, match.start() - 100)
                end = min(len(content), match.end() + 500)
                context = content[start:end]

                # Extract phone numbers from context
                phone_match = re.search(r'(\d{4,5}\s?\d{3}\s?\d{3,4})', context)
                if phone_match and not prospect_lookup[norm].get('phone'):
                    prospect_lookup[norm]['phone'] = phone_match.group(1)

                # Extract emails from context
                email_match = re.search(r'[\w.+-]+@[\w.-]+\.\w+', context)
                if email_match:
                    email = email_match.group(0)
                    p = prospect_lookup[norm]
                    if not p.get('contact_email') and not p.get('generic_email'):
                        if p.get('contact_name'):
                            p['contact_email'] = email
                        else:
                            p['generic_email'] = email

                # Extract website from context
                web_match = re.search(r'www\.[\w.-]+\.\w+', context)
                if web_match and not prospect_lookup[norm].get('website'):
                    prospect_lookup[norm]['website'] = f'https://{web_match.group(0)}'


def main():
    print('Parsing care homes XLSX...')
    parse_carehomes_xlsx()
    print(f'  -> {len(prospects)} prospects so far')

    print('Parsing housing associations XLSX...')
    parse_housing_associations_xlsx()
    print(f'  -> {len(prospects)} prospects so far')

    print('Parsing offices XLSX...')
    parse_offices_xlsx()
    print(f'  -> {len(prospects)} prospects so far')

    print('Parsing estate agents & developers XLSX...')
    parse_estate_agents_xlsx()
    print(f'  -> {len(prospects)} prospects so far')

    print('Enriching from MD files...')
    parse_md_enrichment()

    # Stats
    categories = {}
    for p in prospects:
        cat = p.get('category', 'Other')
        categories[cat] = categories.get(cat, 0) + 1

    print(f'\nTotal unique prospects: {len(prospects)}')
    print('By category:')
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        print(f'  {cat}: {count}')

    # Write output
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(prospects, f, ensure_ascii=False, indent=2)
    print(f'\nWritten to: {OUTPUT}')


if __name__ == '__main__':
    main()
