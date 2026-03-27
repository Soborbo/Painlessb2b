"""
Enrich existing prospects with geocoding (lat/lng) and missing phone numbers.
Sources: MD reference files + Nominatim geocoding API.
"""

import json
import os
import re
import sys
import time
import urllib.request
import urllib.parse

sys.stdout.reconfigure(encoding='utf-8')

BASE_MD = os.path.join(os.environ.get('TEMP', '/tmp'), 'b2b-ref')
API_BASE = 'http://localhost:4322'
COOKIE = None

# Known city center coordinates for fallback geocoding
CITY_COORDS = {
    'bristol': (51.4545, -2.5879),
    'bath': (51.3811, -2.3590),
    'cardiff': (51.4816, -3.1791),
    'newport': (51.5842, -2.9977),
    'clevedon': (51.4382, -2.8544),
    'portishead': (51.4836, -2.7697),
    'keynsham': (51.4133, -2.4967),
    'yate': (51.5407, -2.4110),
    'thornbury': (51.6118, -2.5227),
    'weston-super-mare': (51.3460, -2.9770),
    'filton': (51.5057, -2.5712),
    'kingswood': (51.4547, -2.5048),
    'midsomer norton': (51.2849, -2.4815),
    'frome': (51.2279, -2.3215),
    'taunton': (51.0154, -3.1000),
    'exeter': (50.7184, -3.5339),
    'gloucester': (51.8642, -2.2382),
    'swindon': (51.5558, -1.7797),
    'bridgend': (51.5043, -3.5764),
    'caerphilly': (51.5736, -3.2181),
    'cwmbran': (51.6531, -3.0201),
    'penarth': (51.4345, -3.1716),
    'llanishen': (51.5220, -3.1836),
    'pontypridd': (51.6016, -3.3428),
    'swansea': (51.6214, -3.9436),
    'merthyr tydfil': (51.7459, -3.3786),
    'torfaen': (51.7040, -3.0440),
    'monmouthshire': (51.8116, -2.7163),
    'emersons green': (51.4937, -2.4917),
    'stoke gifford': (51.5139, -2.5448),
    'westbury-on-trym': (51.4892, -2.6103),
    'horfield': (51.4816, -2.5878),
    'bedminster': (51.4380, -2.5945),
    'clifton': (51.4553, -2.6200),
    'fishponds': (51.4753, -2.5248),
    'southville': (51.4410, -2.5973),
    'knowle': (51.4285, -2.5721),
    'whitehall': (51.4643, -2.5567),
    'bishopsworth': (51.4188, -2.6082),
    'hengrove': (51.4124, -2.5794),
    'stockwood': (51.4076, -2.5507),
    'radyr': (51.5108, -3.2541),
    'llanwern': (51.5705, -2.9469),
    'failand': (51.4530, -2.6891),
    'limpley stoke': (51.3540, -2.3200),
    'sandford': (51.3341, -2.8229),
    'chew valley': (51.3400, -2.6200),
    'portishead': (51.4836, -2.7697),
    'almondsbury': (51.5490, -2.5710),
    'patchway': (51.5182, -2.5817),
    'southmead': (51.4988, -2.5890),
    'downend': (51.4850, -2.4997),
    'st george': (51.4578, -2.5385),
    'winterbourne': (51.5270, -2.4843),
    'yatton': (51.3882, -2.8275),
    'worle': (51.3527, -2.9232),
    'nailsea': (51.4306, -2.7573),
    'bradley stoke': (51.5296, -2.5566),
    'hampshire': (51.0577, -1.3081),
    'somerset': (51.1053, -2.9263),
    'wiltshire': (51.3492, -1.9927),
    'cornwall': (50.2660, -5.0527),
    'devon': (50.7156, -3.5309),
    'dorset': (50.7488, -2.3445),
    'south gloucestershire': (51.5250, -2.4800),
    'north somerset': (51.3890, -2.7780),
}

# UK postcode area center coordinates (first part of postcode)
POSTCODE_AREAS = {
    'BS1': (51.4530, -2.5970), 'BS2': (51.4570, -2.5800),
    'BS3': (51.4350, -2.5950), 'BS4': (51.4250, -2.5550),
    'BS5': (51.4600, -2.5500), 'BS6': (51.4700, -2.6000),
    'BS7': (51.4780, -2.5900), 'BS8': (51.4570, -2.6200),
    'BS9': (51.4890, -2.6150), 'BS10': (51.5060, -2.6000),
    'BS11': (51.4950, -2.6800), 'BS13': (51.4130, -2.6200),
    'BS14': (51.4080, -2.5550), 'BS15': (51.4530, -2.5050),
    'BS16': (51.4950, -2.4950), 'BS20': (51.4830, -2.7700),
    'BS21': (51.4380, -2.8550), 'BS22': (51.3460, -2.9360),
    'BS23': (51.3500, -2.9750), 'BS24': (51.3300, -2.9500),
    'BS25': (51.3200, -2.8300), 'BS29': (51.3400, -2.8300),
    'BS30': (51.4400, -2.4700), 'BS31': (51.4130, -2.4970),
    'BS32': (51.5300, -2.5570), 'BS34': (51.5280, -2.5500),
    'BS35': (51.5800, -2.5300), 'BS36': (51.5050, -2.4850),
    'BS37': (51.5400, -2.4100), 'BS39': (51.3200, -2.5500),
    'BS40': (51.3400, -2.7000), 'BS41': (51.4100, -2.6500),
    'BS48': (51.4300, -2.7600), 'BS49': (51.3600, -2.8100),
    'BA1': (51.3900, -2.3600), 'BA2': (51.3700, -2.3600),
    'BA3': (51.2900, -2.4800), 'BA5': (51.2100, -2.6500),
    'CF5': (51.4800, -3.2300), 'CF10': (51.4730, -3.1700),
    'CF11': (51.4670, -3.1900), 'CF14': (51.5150, -3.2000),
    'CF23': (51.5100, -3.1500), 'CF24': (51.4870, -3.1600),
    'NP10': (51.5800, -3.0200), 'NP20': (51.5850, -3.0000),
    'NP44': (51.6530, -3.0200),
}


def get_cookie():
    """Login and get session cookie."""
    global COOKIE
    data = json.dumps({'password': 'testpass123'}).encode()
    req = urllib.request.Request(
        f'{API_BASE}/api/auth/login',
        data=data,
        headers={'Content-Type': 'application/json'},
        method='POST'
    )
    resp = urllib.request.urlopen(req)
    for header in resp.headers.get_all('Set-Cookie') or []:
        if 'session=' in header:
            COOKIE = header.split(';')[0]
            break
    print(f'Logged in: {COOKIE[:30]}...')


def api_get(path):
    """GET from API."""
    req = urllib.request.Request(
        f'{API_BASE}{path}',
        headers={'Cookie': COOKIE}
    )
    resp = urllib.request.urlopen(req)
    return json.loads(resp.read().decode())


def api_put(path, data):
    """PUT to API."""
    body = json.dumps(data).encode()
    req = urllib.request.Request(
        f'{API_BASE}{path}',
        data=body,
        headers={'Content-Type': 'application/json', 'Cookie': COOKIE},
        method='PUT'
    )
    try:
        resp = urllib.request.urlopen(req)
        return json.loads(resp.read().decode())
    except Exception as e:
        print(f'  PUT error: {e}')
        return None


def extract_company_data_from_md():
    """Parse MD files to extract postcodes, addresses, and phone numbers per company."""
    company_data = {}  # normalized name -> {postcode, address, phone, ...}

    md_files = [
        'Later-life-prospect-pipeline.md',
        'housing-associations-pipeline.md',
        'commercial-office-prospects-pipeline.md',
    ]

    for fname in md_files:
        path = os.path.join(BASE_MD, fname)
        if not os.path.exists(path):
            continue
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Split into sections by ### headers or table rows with bold company names
        # Strategy: find each **CompanyName** and extract data from surrounding context

        # First approach: parse the detailed Field|Detail tables
        # These look like: | **Type** | detail |
        sections = re.split(r'### ', content)
        for section in sections:
            # Get company name from section header
            header_match = re.match(r'([^\n]+)', section)
            if not header_match:
                continue
            header = header_match.group(1).strip()
            # Clean header: remove " — description" suffix
            company_name = re.sub(r'\s*[—–-]\s+.*$', '', header).strip()
            if not company_name or len(company_name) < 3:
                continue

            norm = re.sub(r'\s+', ' ', company_name.lower().strip())

            data = {}

            # Extract postcode from section
            pc_matches = re.findall(r'([A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2})', section)
            if pc_matches:
                data['postcode'] = pc_matches[0]  # Take first postcode as primary

            # Extract phone from section (UK format)
            phone_matches = re.findall(r'(0\d{2,4}[\s-]?\d{3}[\s-]?\d{3,4})', section)
            if phone_matches:
                data['phone'] = phone_matches[0]

            # Extract specific address from | **Address** | or location info
            addr_match = re.search(r'HQ[:\s]+([^|;\n]+)', section)
            if addr_match:
                data['hq_address'] = addr_match.group(1).strip()

            if data:
                company_data[norm] = data

        # Second approach: parse markdown table rows
        # | N | **Name** | Type | ... |
        for line in content.split('\n'):
            if '|' not in line or '---' in line:
                continue
            cells = [c.strip() for c in line.split('|')]
            # Find the cell with bold company name
            for i, cell in enumerate(cells):
                bold_match = re.match(r'\*\*([^*]+)\*\*', cell)
                if bold_match:
                    company_name = bold_match.group(1).strip()
                    norm = re.sub(r'\s+', ' ', company_name.lower().strip())

                    if norm not in company_data:
                        company_data[norm] = {}

                    # Extract postcode from the full row
                    row_text = line
                    pc_matches = re.findall(r'([A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2})', row_text)
                    if pc_matches and 'postcode' not in company_data[norm]:
                        company_data[norm]['postcode'] = pc_matches[0]

                    # Extract phone from row
                    phone_matches = re.findall(r'(0\d{2,4}[\s-]?\d{3}[\s-]?\d{3,4})', row_text)
                    if phone_matches and 'phone' not in company_data[norm]:
                        company_data[norm]['phone'] = phone_matches[0]
                    break

    # Also parse XLSX extra sheets for more data
    try:
        import openpyxl
        # Building Map sheet has addresses
        xlsx_path = os.path.join(BASE_MD, 'painless_removals_offices_pipeline.xlsx')
        wb = openpyxl.load_workbook(xlsx_path)
        if 'Building Map' in wb.sheetnames:
            ws = wb['Building Map']
            headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=False):
                vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}
                agent = str(vals.get('Managing Agent / Letting Agent', '') or '')
                # Extract managing agents
                for agent_name in re.split(r'[/,;]', agent):
                    agent_name = agent_name.strip().strip('()')
                    if len(agent_name) > 3:
                        norm = re.sub(r'\s+', ' ', agent_name.lower().strip())
                        addr = str(vals.get('Address', '') or '')
                        if addr and norm not in company_data:
                            company_data[norm] = {}
                        if addr:
                            pc_matches = re.findall(r'([A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2})', addr)
                            if pc_matches and 'postcode' not in company_data.get(norm, {}):
                                company_data.setdefault(norm, {})['postcode'] = pc_matches[0]

        # Development Sites sheet
        xlsx_path2 = os.path.join(BASE_MD, 'painless_removals_estate_agents_developers_pipeline.xlsx')
        wb2 = openpyxl.load_workbook(xlsx_path2)
        if 'Development Sites' in wb2.sheetnames:
            ws2 = wb2['Development Sites']
            headers2 = [str(c.value).strip() if c.value else '' for c in ws2[1]]
            for row in ws2.iter_rows(min_row=2, values_only=False):
                vals = {headers2[i]: row[i].value for i in range(len(headers2)) if i < len(row)}
                devs = str(vals.get('Developer(s)', '') or '')
                phone = str(vals.get('Sales Office Phone', '') or '')
                for dev_name in re.split(r'[,;()]', devs):
                    dev_name = dev_name.strip()
                    if len(dev_name) > 3 and dev_name.lower() not in ('lead', 'master', 'via', 'also'):
                        norm = re.sub(r'\s+', ' ', dev_name.lower().strip())
                        if phone and re.match(r'0\d', phone):
                            company_data.setdefault(norm, {})['phone'] = phone

    except Exception as e:
        print(f'  XLSX enrichment error: {e}')

    return company_data


def geocode_nominatim(query):
    """Geocode using Nominatim API. Returns (lat, lng) or None."""
    params = urllib.parse.urlencode({
        'q': query,
        'format': 'json',
        'limit': 1,
        'countrycodes': 'gb',
    })
    url = f'https://nominatim.openstreetmap.org/search?{params}'
    req = urllib.request.Request(url, headers={
        'User-Agent': 'PainlessRemovalsBulkImport/1.0'
    })
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode())
        if data:
            return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        pass
    return None


def geocode_prospect(company):
    """Try to geocode a prospect using best available info."""
    name = company.get('name', '')
    address = company.get('address', '') or ''
    postcode = company.get('postcode', '') or ''

    # Strategy 1: Use postcode (most accurate)
    if postcode:
        # Try postcode area lookup first (no API call needed)
        pc_area = re.match(r'([A-Z]{1,2}\d{1,2})', postcode)
        if pc_area and pc_area.group(1) in POSTCODE_AREAS:
            return POSTCODE_AREAS[pc_area.group(1)]
        # Full postcode via Nominatim
        result = geocode_nominatim(postcode)
        if result:
            return result
        time.sleep(1.1)  # Rate limit

    # Strategy 2: Use address with company name
    if address:
        # Try to match city from address
        addr_lower = address.lower()
        for city, coords in CITY_COORDS.items():
            if city in addr_lower:
                # Add some jitter to avoid all pins stacking
                import random
                jitter_lat = random.uniform(-0.008, 0.008)
                jitter_lng = random.uniform(-0.008, 0.008)
                return (coords[0] + jitter_lat, coords[1] + jitter_lng)

    # Strategy 3: Use website domain to guess location (skip - unreliable)

    # Strategy 4: Default to Bristol center with jitter
    import random
    jitter_lat = random.uniform(-0.015, 0.015)
    jitter_lng = random.uniform(-0.015, 0.015)
    return (51.4545 + jitter_lat, -2.5879 + jitter_lng)


def find_best_match(norm_name, md_data):
    """Find best match for a company name in MD data."""
    # Exact match
    if norm_name in md_data:
        return md_data[norm_name]

    # Partial match - company name contains or is contained
    for md_name, data in md_data.items():
        if md_name in norm_name or norm_name in md_name:
            return data
        # Match on first significant word
        words_a = [w for w in norm_name.split() if len(w) > 3]
        words_b = [w for w in md_name.split() if len(w) > 3]
        if words_a and words_b and words_a[0] == words_b[0]:
            return data

    return None


def main():
    import random
    random.seed(42)  # Reproducible jitter

    print('Step 1: Extracting data from MD files...')
    md_data = extract_company_data_from_md()
    print(f'  Found data for {len(md_data)} companies in MD files')
    phones_found = sum(1 for d in md_data.values() if 'phone' in d)
    postcodes_found = sum(1 for d in md_data.values() if 'postcode' in d)
    print(f'  Phones: {phones_found}, Postcodes: {postcodes_found}')

    print('\nStep 2: Logging in...')
    get_cookie()

    print('\nStep 3: Fetching all companies...')
    companies = api_get('/api/companies')
    print(f'  Got {len(companies)} companies')

    no_phone = sum(1 for c in companies if not c.get('phone'))
    no_coords = sum(1 for c in companies if not c.get('lat'))
    print(f'  Missing phone: {no_phone}')
    print(f'  Missing lat/lng: {no_coords}')

    print('\nStep 4: Enriching and geocoding...')
    updated = 0
    phones_added = 0
    geocoded = 0
    nominatim_calls = 0

    for i, company in enumerate(companies):
        norm = re.sub(r'\s+', ' ', company['name'].lower().strip())
        update_data = {}

        # Find MD data match
        md_match = find_best_match(norm, md_data)

        # Add postcode if missing
        if not company.get('postcode') and md_match and md_match.get('postcode'):
            update_data['postcode'] = md_match['postcode']

        # Add phone if missing
        if not company.get('phone') and md_match and md_match.get('phone'):
            update_data['phone'] = md_match['phone']
            phones_added += 1

        # Geocode if missing
        if not company.get('lat'):
            # Use postcode from update or existing
            pc = update_data.get('postcode') or company.get('postcode')
            geocode_input = dict(company)
            if pc:
                geocode_input['postcode'] = pc

            coords = geocode_prospect(geocode_input)
            if coords:
                update_data['lat'] = round(coords[0], 6)
                update_data['lng'] = round(coords[1], 6)
                geocoded += 1

        if update_data:
            result = api_put(f'/api/companies/{company["id"]}', update_data)
            if result:
                updated += 1

        if (i + 1) % 20 == 0:
            print(f'  Processed {i + 1}/{len(companies)}...')

    print(f'\nDone!')
    print(f'  Companies updated: {updated}')
    print(f'  Phones added: {phones_added}')
    print(f'  Geocoded: {geocoded}')


if __name__ == '__main__':
    main()
